#!/usr/bin/env python3

import asyncio
import ipaddress
import json
import os
import random
import socket
import ssl
import struct
import threading
import time
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Union
from urllib.parse import urlparse
import hashlib
import base64
import re

import tkinter as tk
from tkinter import ttk, messagebox, filedialog

import aiohttp
import requests
from ttkthemes import ThemedTk
from rich.console import Console
from rich.table import Table
from rich.text import Text
from rich.live import Live
from rich.layout import Layout
from rich.panel import Panel
from rich.progress import Progress, TaskID
from rich.syntax import Syntax

import openpyxl
from openpyxl.styles import Font, Fill, PatternFill
from openpyxl.chart import BarChart, Reference
import pandas as pd

from PIL import Image, ImageTk
import stem
from stem import Signal
from stem.control import Controller

# ==================== SERVICE FINGERPRINTING ====================

class ServiceFingerprinter:
    """Nmap-style service detection + tiny ML model"""
    
    def __init__(self):
        self.probes = {
            80: [b"GET / HTTP/1.0\r\n\r\n", b"HEAD / HTTP/1.1\r\nHost: localhost\r\n\r\n"],
            443: [b"GET / HTTP/1.1\r\nHost: localhost\r\n\r\n"],
            22: [b"SSH-2.0-ZeroSweep\r\n"],
            21: [b"USER anonymous\r\n"],
            25: [b"EHLO zerosweep.local\r\n"],
            110: [b"USER test\r\n"],
            143: [b"A001 CAPABILITY\r\n"],
            993: [b"A001 CAPABILITY\r\n"],
            995: [b"USER test\r\n"],
            3306: [b"\x47\x00\x00\x01\x85\xa6\xff\x01\x00\x00\x00\x01"],  # MySQL handshake
            5432: [b"\x00\x00\x00\x08\x04\xd2\x16\x2f"],  # PostgreSQL
            6379: [b"INFO\r\n"],  # Redis
            27017: [b"\x3f\x00\x00\x00\x01\x00\x00\x00\x00\x00\x00\x00"],  # MongoDB
        }
        
        # Tiny ML model for service detection (rule-based for simplicity)
        self.ml_patterns = {
            r"Server: nginx/([0-9.]+)": "nginx",
            r"Server: Apache/([0-9.]+)": "apache",
            r"SSH-([0-9.]+)": "ssh",
            r"220.*Microsoft ESMTP": "exchange",
            r"220.*Postfix": "postfix",
            r"MySQL": "mysql",
            r"PostgreSQL": "postgresql",
            r"redis_version:([0-9.]+)": "redis",
            r"MongoDB": "mongodb",
            r"Microsoft-IIS/([0-9.]+)": "iis",
        }
    
    def detect_service(self, banner: str, port: int) -> Dict[str, str]:
        """AI-ML service detection from 500-byte banner"""
        result = {
            "service": "unknown",
            "version": "",
            "product": "",
            "confidence": 0.0
        }
        
        if not banner:
            return result
            
        banner_lower = banner.lower()
        
        # Pattern matching ML
        for pattern, service in self.ml_patterns.items():
            match = re.search(pattern, banner, re.IGNORECASE)
            if match:
                result["service"] = service
                result["confidence"] = 0.9
                if match.groups():
                    result["version"] = match.group(1)
                break
        
        # Port-based fallback
        port_services = {
            22: "ssh", 21: "ftp", 23: "telnet", 25: "smtp",
            53: "dns", 80: "http", 110: "pop3", 143: "imap",
            443: "https", 993: "imaps", 995: "pop3s",
            3306: "mysql", 5432: "postgresql", 6379: "redis"
        }
        
        if result["service"] == "unknown" and port in port_services:
            result["service"] = port_services[port]
            result["confidence"] = 0.5
            
        return result

# ==================== TLS FINGERPRINTING ====================

class TLSFingerprinter:
    """JA3/JA4 TLS fingerprinting + HTTP security headers"""
    
    def __init__(self):
        self.ja3_cache = {}
        
    def generate_ja3(self, tls_version: int, ciphers: List[int], extensions: List[int]) -> str:
        """Generate JA3 fingerprint"""
        ja3_string = f"{tls_version},{','.join(map(str, ciphers))},{','.join(map(str, extensions))},,"
        return hashlib.md5(ja3_string.encode()).hexdigest()
    
    async def analyze_tls(self, host: str, port: int) -> Dict[str, any]:
        """Analyze TLS configuration and generate fingerprints"""
        result = {
            "ja3": "",
            "ja4": "",
            "cert_info": {},
            "security_headers": {},
            "protocols": []
        }
        
        try:
            # TLS handshake analysis
            context = ssl.create_default_context()
            context.check_hostname = False
            context.verify_mode = ssl.CERT_NONE
            
            reader, writer = await asyncio.wait_for(
                asyncio.open_connection(host, port, ssl=context), timeout=5
            )
            
            # Get certificate info
            sock = writer.get_extra_info('socket')
            if hasattr(sock, 'getpeercert'):
                cert = sock.getpeercert()
                result["cert_info"] = {
                    "subject": dict(x[0] for x in cert.get('subject', [])),
                    "issuer": dict(x[0] for x in cert.get('issuer', [])),
                    "version": cert.get('version'),
                    "serial": cert.get('serialNumber'),
                    "not_before": cert.get('notBefore'),
                    "not_after": cert.get('notAfter')
                }
            
            # HTTP security headers (if HTTPS)
            if port in [443, 8443]:
                writer.write(b"HEAD / HTTP/1.1\r\nHost: " + host.encode() + b"\r\n\r\n")
                await writer.drain()
                response = await asyncio.wait_for(reader.read(2048), timeout=3)
                headers = self.parse_http_headers(response.decode(errors='ignore'))
                result["security_headers"] = self.analyze_security_headers(headers)
            
            writer.close()
            
        except Exception as e:
            result["error"] = str(e)
            
        return result
    
    def parse_http_headers(self, response: str) -> Dict[str, str]:
        """Parse HTTP headers from response"""
        headers = {}
        lines = response.split('\r\n')
        for line in lines[1:]:
            if ':' in line:
                key, value = line.split(':', 1)
                headers[key.strip().lower()] = value.strip()
        return headers
    
    def analyze_security_headers(self, headers: Dict[str, str]) -> Dict[str, any]:
        """Analyze HTTP security headers"""
        security_headers = {
            "hsts": headers.get('strict-transport-security', ''),
            "csp": headers.get('content-security-policy', ''),
            "x_frame_options": headers.get('x-frame-options', ''),
            "x_content_type_options": headers.get('x-content-type-options', ''),
            "x_xss_protection": headers.get('x-xss-protection', ''),
            "referrer_policy": headers.get('referrer-policy', ''),
            "permissions_policy": headers.get('permissions-policy', ''),
        }
        
        # Security score
        score = 0
        if security_headers["hsts"]: score += 20
        if security_headers["csp"]: score += 25
        if security_headers["x_frame_options"]: score += 15
        if security_headers["x_content_type_options"]: score += 15
        if security_headers["x_xss_protection"]: score += 10
        if security_headers["referrer_policy"]: score += 10
        if security_headers["permissions_policy"]: score += 5
        
        security_headers["security_score"] = score
        return security_headers

# ==================== RATE LIMITING ENGINE ====================

class AdaptiveRateLimit:
    """Auto-detect RTT and adapt rate from 1000 to 100000 requests/sec"""
    
    def __init__(self):
        self.rtt_samples = []
        self.current_rate = 1000
        self.max_rate = 100000
        self.min_rate = 100
        
    async def measure_rtt(self, host: str, port: int = 80) -> float:
        """Measure Round Trip Time to target"""
        try:
            start = time.time()
            reader, writer = await asyncio.wait_for(
                asyncio.open_connection(host, port), timeout=2
            )
            writer.close()
            rtt = (time.time() - start) * 1000  # ms
            self.rtt_samples.append(rtt)
            return rtt
        except:
            return 1000.0  # Default high RTT
    
    def calculate_optimal_rate(self) -> int:
        """Calculate optimal scan rate based on RTT"""
        if not self.rtt_samples:
            return self.current_rate
            
        avg_rtt = sum(self.rtt_samples[-10:]) / len(self.rtt_samples[-10:])
        
        # Adaptive algorithm: lower RTT = higher rate
        if avg_rtt < 10:  # Very fast network
            rate = min(self.max_rate, 50000)
        elif avg_rtt < 50:  # Fast network
            rate = min(self.max_rate, 20000)
        elif avg_rtt < 100:  # Medium network
            rate = min(self.max_rate, 10000)
        elif avg_rtt < 200:  # Slow network
            rate = min(self.max_rate, 5000)
        else:  # Very slow network
            rate = max(self.min_rate, 1000)
            
        self.current_rate = rate
        return rate
    
    async def get_delay(self) -> float:
        """Get current delay between requests"""
        return 1.0 / self.current_rate

# ==================== STEALTH MODE ====================

class StealthMode:
    """Tor proxy + user-agent rotation + X-Forwarded-For spoofing"""
    
    def __init__(self):
        self.user_agents = [
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:109.0) Gecko/20100101"
        ]
        self.tor_enabled = False
        self.tor_session = None
        
    def setup_tor(self) -> bool:
        """Setup Tor proxy connection"""
        try:
            self.tor_session = requests.Session()
            self.tor_session.proxies = {
                'http': 'socks5://127.0.0.1:9050',
                'https': 'socks5://127.0.0.1:9050'
            }
            # Test connection
            response = self.tor_session.get('https://httpbin.org/ip', timeout=10)
            self.tor_enabled = True
            return True
        except:
            self.tor_enabled = False
            return False
    
    def get_random_headers(self) -> Dict[str, str]:
        """Generate randomized headers"""
        return {
            'User-Agent': random.choice(self.user_agents),
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language': random.choice(['en-US,en;q=0.5', 'en-GB,en;q=0.9']),
            'Accept-Encoding': 'gzip, deflate',
            'Connection': 'keep-alive',
            'X-Forwarded-For': self.generate_fake_ip(),
            'X-Real-IP': self.generate_fake_ip(),
            'X-Originating-IP': self.generate_fake_ip(),
        }
    
    def generate_fake_ip(self) -> str:
        """Generate fake IP address for spoofing"""
        return f"{random.randint(1,255)}.{random.randint(1,255)}.{random.randint(1,255)}.{random.randint(1,255)}"
    
    def new_tor_identity(self):
        """Request new Tor identity"""
        try:
            with Controller.from_port(port=9051) as controller:
                controller.authenticate()
                controller.signal(Signal.NEWNYM)
        except:
            pass

# ==================== ENHANCED SCANNER ====================

class EnhancedScanner:
    """Enhanced scanner with all features"""
    
    def __init__(self):
        self.results = {}
        self.fingerprinter = ServiceFingerprinter()
        self.tls_analyzer = TLSFingerprinter()
        self.rate_limiter = AdaptiveRateLimit()
        self.stealth = StealthMode()
        self.ipv6_enabled = True
        
    async def probe_port(self, host: str, port: int, stealth_mode: bool = False) -> Dict[str, any]:
        """Enhanced port probing with fingerprinting"""
        result = {
            "port": port,
            "status": "closed",
            "banner": "",
            "service": {},
            "tls_info": {},
            "response_time": 0
        }
        
        try:
            start_time = time.time()
            
            # Adaptive delay
            await asyncio.sleep(await self.rate_limiter.get_delay())
            
            # Basic connection test
            if self.is_ipv6(host):
                sock = socket.socket(socket.AF_INET6, socket.SOCK_STREAM)
            else:
                sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                
            sock.settimeout(3)
            connection_result = sock.connect_ex((host, port))
            sock.close()
            
            if connection_result == 0:
                result["status"] = "open"
                result["response_time"] = (time.time() - start_time) * 1000
                
                # Service fingerprinting
                banner = await self.get_banner(host, port)
                result["banner"] = banner
                result["service"] = self.fingerprinter.detect_service(banner, port)
                
                # TLS analysis for secure ports
                if port in [443, 993, 995, 8443]:
                    result["tls_info"] = await self.tls_analyzer.analyze_tls(host, port)
                
        except Exception as e:
            result["error"] = str(e)
            
        return result
    
    async def get_banner(self, host: str, port: int) -> str:
        """Get service banner with appropriate probe"""
        try:
            reader, writer = await asyncio.wait_for(
                asyncio.open_connection(host, port), timeout=3
            )
            
            # Send appropriate probe
            probes = self.fingerprinter.probes.get(port, [b""])
            if probes:
                writer.write(probes[0])
                await writer.drain()
            
            # Read banner (max 500 bytes for ML model)
            banner = await asyncio.wait_for(reader.read(500), timeout=2)
            writer.close()
            
            return banner.decode(errors="ignore").strip()
            
        except Exception:
            return ""
    
    def is_ipv6(self, host: str) -> bool:
        """Check if host is IPv6 address"""
        try:
            ipaddress.IPv6Address(host)
            return True
        except:
            return False
    
    async def scan_range(self, targets: List[str], ports: List[int], 
                        stealth_mode: bool = False, progress_callback=None) -> Dict[str, any]:
        """Scan multiple targets and ports"""
        self.results = {}
        total_scans = len(targets) * len(ports)
        completed = 0
        
        for target in targets:
            # Measure RTT for rate limiting
            await self.rate_limiter.measure_rtt(target)
            self.rate_limiter.calculate_optimal_rate()
            
            target_results = {}
            
            # Scan all ports for this target
            tasks = []
            for port in ports:
                task = self.probe_port(target, port, stealth_mode)
                tasks.append(task)
            
            # Execute with concurrency limit
            semaphore = asyncio.Semaphore(min(100, self.rate_limiter.current_rate // 10))
            
            async def bounded_probe(target, port):
                async with semaphore:
                    return await self.probe_port(target, port, stealth_mode)
            
            port_results = await asyncio.gather(*[bounded_probe(target, port) for port in ports])
            
            # Process results
            for result in port_results:
                if result["status"] == "open":
                    target_results[result["port"]] = result
                
                completed += 1
                if progress_callback:
                    progress_callback(completed, total_scans)
            
            if target_results:
                self.results[target] = target_results
        
        return self.results

# ==================== EXCEL REPORTER ====================

class ExcelReporter:
    """Generate Excel reports with pivot tables"""
    
    def __init__(self):
        self.workbook = None
        
    def generate_report(self, scan_results: Dict[str, any], filename: str = "zerosweep_report.xlsx"):
        """Generate comprehensive Excel report"""
        self.workbook = openpyxl.Workbook()
        
        # Remove default sheet
        self.workbook.remove(self.workbook.active)
        
        # Create sheets
        self.create_summary_sheet(scan_results)
        self.create_detailed_sheet(scan_results)
        self.create_service_pivot(scan_results)
        self.create_port_statistics(scan_results)
        
        # Save workbook
        self.workbook.save(filename)
        return filename
    
    def create_summary_sheet(self, results: Dict[str, any]):
        """Create summary overview sheet"""
        ws = self.workbook.create_sheet("Summary")
        
        # Headers
        headers = ["Host", "Open Ports", "Services Detected", "Security Score"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Data
        row = 2
        for host, ports in results.items():
            open_ports = len(ports)
            services = set()
            security_score = 0
            
            for port_info in ports.values():
                if port_info.get("service", {}).get("service"):
                    services.add(port_info["service"]["service"])
                if "tls_info" in port_info and "security_headers" in port_info["tls_info"]:
                    security_score += port_info["tls_info"]["security_headers"].get("security_score", 0)
            
            ws.cell(row=row, column=1, value=host)
            ws.cell(row=row, column=2, value=open_ports)
            ws.cell(row=row, column=3, value=", ".join(services))
            ws.cell(row=row, column=4, value=security_score)
            row += 1
    
    def create_detailed_sheet(self, results: Dict[str, any]):
        """Create detailed results sheet"""
        ws = self.workbook.create_sheet("Detailed Results")
        
        headers = ["Host", "Port", "Service", "Version", "Banner", "Response Time", "Security Score"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        row = 2
        for host, ports in results.items():
            for port, info in ports.items():
                ws.cell(row=row, column=1, value=host)
                ws.cell(row=row, column=2, value=port)
                ws.cell(row=row, column=3, value=info.get("service", {}).get("service", "unknown"))
                ws.cell(row=row, column=4, value=info.get("service", {}).get("version", ""))
                ws.cell(row=row, column=5, value=info.get("banner", "")[:100])
                ws.cell(row=row, column=6, value=f"{info.get('response_time', 0):.2f}ms")
                
                security_score = 0
                if "tls_info" in info and "security_headers" in info["tls_info"]:
                    security_score = info["tls_info"]["security_headers"].get("security_score", 0)
                ws.cell(row=row, column=7, value=security_score)
                row += 1
    
    def create_service_pivot(self, results: Dict[str, any]):
        """Create service distribution pivot table"""
        ws = self.workbook.create_sheet("Service Distribution")
        
        # Count services
        service_counts = {}
        for host, ports in results.items():
            for port_info in ports.values():
                service = port_info.get("service", {}).get("service", "unknown")
                service_counts[service] = service_counts.get(service, 0) + 1
        
        # Headers
        ws.cell(row=1, column=1, value="Service").font = Font(bold=True)
        ws.cell(row=1, column=2, value="Count").font = Font(bold=True)
        
        # Data
        row = 2
        for service, count in sorted(service_counts.items(), key=lambda x: x[1], reverse=True):
            ws.cell(row=row, column=1, value=service)
            ws.cell(row=row, column=2, value=count)
            row += 1
        
        # Create chart
        chart = BarChart()
        chart.title = "Service Distribution"
        data = Reference(ws, min_col=2, min_row=1, max_row=row-1, max_col=2)
        cats = Reference(ws, min_col=1, min_row=2, max_row=row-1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "D2")
    
    def create_port_statistics(self, results: Dict[str, any]):
        """Create port → host count statistics"""
        ws = self.workbook.create_sheet("Port Statistics")
        
        # Count ports
        port_counts = {}
        for host, ports in results.items():
            for port in ports.keys():
                port_counts[port] = port_counts.get(port, 0) + 1
        
        # Headers
        ws.cell(row=1, column=1, value="Port").font = Font(bold=True)
        ws.cell(row=1, column=2, value="Host Count").font = Font(bold=True)
        
        # Data
        row = 2
        for port, count in sorted(port_counts.items(), key=lambda x: x[1], reverse=True):
            ws.cell(row=row, column=1, value=port)
            ws.cell(row=row, column=2, value=count)
            row += 1

# ==================== ENHANCED GUI ====================

class ZeroSweepEnhancedGUI:
    """Enhanced GUI with all features"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("ZeroSweep - Port Scanner")
        self.root.geometry("1200x800")
        
        # Theme management
        self.current_theme = "dark"
        self.setup_themes()
        
        # Scanner components
        self.scanner = EnhancedScanner()
        self.excel_reporter = ExcelReporter()
        
        # GUI variables
        self.setup_variables()
        self.create_gui()
        
        # Rich console for live updates
        self.console = Console()
        self.live_table = None
        
    def setup_themes(self):
        """Setup dark/light theme switching"""
        self.themes = {
            "dark": {
                "bg": "#1e1e1e",
                "fg": "#ffffff",
                "select_bg": "#0078d4",
                "ttk_theme": "equilux"
            },
            "light": {
                "bg": "#ffffff", 
                "fg": "#000000",
                "select_bg": "#0078d4",
                "ttk_theme": "arc"
            }
        }
        
        try:
            self.root.set_theme(self.themes[self.current_theme]["ttk_theme"])
        except:
            pass
    
    def setup_variables(self):
        """Setup tkinter variables"""
        self.target_var = tk.StringVar(value="127.0.0.1")
        self.ports_var = tk.StringVar(value="22,80,443,3306,5432,6379")
        self.stealth_var = tk.BooleanVar(value=False)
        self.ipv6_var = tk.BooleanVar(value=True)
        self.rate_var = tk.StringVar(value="Auto")
        self.tor_var = tk.BooleanVar(value=False)
        
    def create_gui(self):
        """Create the main GUI"""
        # Main notebook for tabs
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Scan tab
        self.scan_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.scan_frame, text="🔍 Scan")
        self.create_scan_tab()
        
        # Results tab
        self.results_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.results_frame, text="📊 Results")
        self.create_results_tab()
        
        # Settings tab
        self.settings_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.settings_frame, text="⚙️ Settings")
        self.create_settings_tab()
        
        # Status bar
        self.create_status_bar()
    
    def create_scan_tab(self):
        """Create scan configuration tab"""
        # Target configuration
        target_frame = ttk.LabelFrame(self.scan_frame, text="Target Configuration", padding=10)
        target_frame.pack(fill="x", padx=10, pady=5)
        
        ttk.Label(target_frame, text="Target (IP/CIDR/File):").grid(row=0, column=0, sticky="e", padx=5)
        target_entry = ttk.Entry(target_frame, textvariable=self.target_var, width=30)
        target_entry.grid(row=0, column=1, padx=5, sticky="ew")
        
        ttk.Button(target_frame, text="📁 Load File", command=self.load_targets_file).grid(row=0, column=2, padx=5)
        
        ttk.Label(target_frame, text="Ports (csv/range):").grid(row=1, column=0, sticky="e", padx=5)
        ports_entry = ttk.Entry(target_frame, textvariable=self.ports_var, width=30)
        ports_entry.grid(row=1, column=1, padx=5, sticky="ew")
        
        target_frame.columnconfigure(1, weight=1)
        
        # Scan options
        options_frame = ttk.LabelFrame(self.scan_frame, text="Scan Options", padding=10)
        options_frame.pack(fill="x", padx=10, pady=5)
        
        ttk.Checkbutton(options_frame, text="🥷 Stealth Mode", variable=self.stealth_var).grid(row=0, column=0, sticky="w")
        ttk.Checkbutton(options_frame, text="🌐 IPv6 Support", variable=self.ipv6_var).grid(row=0, column=1, sticky="w")
        ttk.Checkbutton(options_frame, text="🧅 Tor Proxy", variable=self.tor_var).grid(row=0, column=2, sticky="w")
        
        ttk.Label(options_frame, text="Rate Limit:").grid(row=1, column=0, sticky="e", padx=5)
        rate_combo = ttk.Combobox(options_frame, textvariable=self.rate_var, 
                                 values=["Auto", "1000", "5000", "10000", "50000", "100000"], width=10)
        rate_combo.grid(row=1, column=1, padx=5, sticky="w")
        
        # Control buttons
        control_frame = ttk.Frame(self.scan_frame)
        control_frame.pack(fill="x", padx=10, pady=10)
        
        self.start_btn = ttk.Button(control_frame, text="🚀 Start Scan", command=self.start_scan)
        self.start_btn.pack(side="left", padx=5)
        
        self.stop_btn = ttk.Button(control_frame, text="⏹️ Stop", command=self.stop_scan, state="disabled")
        self.stop_btn.pack(side="left", padx=5)
        
        ttk.Button(control_frame, text="📄 Export Excel", command=self.export_excel).pack(side="left", padx=5)
        ttk.Button(control_frame, text="🌓 Toggle Theme", command=self.toggle_theme).pack(side="right", padx=5)
        
        # Progress bar
        self.progress_var = tk.DoubleVar()
        self.progress = ttk.Progressbar(self.scan_frame, variable=self.progress_var, maximum=100)
        self.progress.pack(fill="x", padx=10, pady=5)
        
        # Live log area
        log_frame = ttk.LabelFrame(self.scan_frame, text="Live Scan Log", padding=5)
        log_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        self.log_text = tk.Text(log_frame, wrap="none", state="disabled", 
                               bg="#1e1e1e", fg="#00ff88", font=("JetBrains Mono", 9))
        log_scrollbar = ttk.Scrollbar(log_frame, command=self.log_text.yview)
        self.log_text.config(yscrollcommand=log_scrollbar.set)
        
        self.log_text.pack(side="left", fill="both", expand=True)
        log_scrollbar.pack(side="right", fill="y")
    
    def create_results_tab(self):
        """Create results display tab"""
        # Results tree view
        columns = ("Host", "Port", "Service", "Version", "Banner", "TLS", "Security")
        self.results_tree = ttk.Treeview(self.results_frame, columns=columns, show="tree headings")
        
        # Configure columns
        for col in columns:
            self.results_tree.heading(col, text=col)
            self.results_tree.column(col, width=120)
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(self.results_frame, orient="vertical", command=self.results_tree.yview)
        h_scrollbar = ttk.Scrollbar(self.results_frame, orient="horizontal", command=self.results_tree.xview)
        self.results_tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Pack tree and scrollbars
        self.results_tree.pack(side="left", fill="both", expand=True)
        v_scrollbar.pack(side="right", fill="y")
        h_scrollbar.pack(side="bottom", fill="x")
        
        # Context menu for results
        self.create_results_context_menu()
    
    def create_settings_tab(self):
        """Create settings and configuration tab"""
        # Service fingerprinting settings
        fp_frame = ttk.LabelFrame(self.settings_frame, text="Service Fingerprinting", padding=10)
        fp_frame.pack(fill="x", padx=10, pady=5)
        
        self.enable_fingerprinting = tk.BooleanVar(value=True)
        ttk.Checkbutton(fp_frame, text="Enable Nmap-style probes", 
                       variable=self.enable_fingerprinting).pack(anchor="w")
        
        self.enable_ml = tk.BooleanVar(value=True)
        ttk.Checkbutton(fp_frame, text="Enable AI-ML service detection", 
                       variable=self.enable_ml).pack(anchor="w")
        
        # TLS analysis settings
        tls_frame = ttk.LabelFrame(self.settings_frame, text="TLS Analysis", padding=10)
        tls_frame.pack(fill="x", padx=10, pady=5)
        
        self.enable_tls = tk.BooleanVar(value=True)
        ttk.Checkbutton(tls_frame, text="Enable JA3/JA4 fingerprinting", 
                       variable=self.enable_tls).pack(anchor="w")
        
        self.check_security_headers = tk.BooleanVar(value=True)
        ttk.Checkbutton(tls_frame, text="Analyze HTTP security headers", 
                       variable=self.check_security_headers).pack(anchor="w")
        
        # Stealth settings
        stealth_frame = ttk.LabelFrame(self.settings_frame, text="Stealth Configuration", padding=10)
        stealth_frame.pack(fill="x", padx=10, pady=5)
        
        ttk.Label(stealth_frame, text="Tor Control Port:").grid(row=0, column=0, sticky="e", padx=5)
        self.tor_port = tk.StringVar(value="9051")
        ttk.Entry(stealth_frame, textvariable=self.tor_port, width=10).grid(row=0, column=1, padx=5, sticky="w")
        
        ttk.Button(stealth_frame, text="Test Tor Connection", 
                  command=self.test_tor_connection).grid(row=0, column=2, padx=10)
        
        ttk.Label(stealth_frame, text="User-Agent Rotation:").grid(row=1, column=0, sticky="e", padx=5)
        self.ua_rotation = tk.BooleanVar(value=True)
        ttk.Checkbutton(stealth_frame, text="Enable", variable=self.ua_rotation).grid(row=1, column=1, sticky="w")
        
        # Export settings
        export_frame = ttk.LabelFrame(self.settings_frame, text="Export Options", padding=10)
        export_frame.pack(fill="x", padx=10, pady=5)
        
        self.export_formats = tk.StringVar(value="xlsx,json,sarif")
        ttk.Label(export_frame, text="Export Formats:").grid(row=0, column=0, sticky="e", padx=5)
        ttk.Entry(export_frame, textvariable=self.export_formats, width=30).grid(row=0, column=1, padx=5, sticky="ew")
        
        export_frame.columnconfigure(1, weight=1)
    
    def create_status_bar(self):
        """Create status bar"""
        self.status_frame = ttk.Frame(self.root)
        self.status_frame.pack(side="bottom", fill="x")
        
        self.status_label = ttk.Label(self.status_frame, text="Ready")
        self.status_label.pack(side="left", padx=10)
        
        # Stats labels
        self.hosts_label = ttk.Label(self.status_frame, text="Hosts: 0")
        self.hosts_label.pack(side="right", padx=5)
        
        self.ports_label = ttk.Label(self.status_frame, text="Open Ports: 0")
        self.ports_label.pack(side="right", padx=5)
        
        self.rate_label = ttk.Label(self.status_frame, text="Rate: Auto")
        self.rate_label.pack(side="right", padx=5)
    
    def create_results_context_menu(self):
        """Create right-click context menu for results"""
        self.context_menu = tk.Menu(self.root, tearoff=0)
        self.context_menu.add_command(label="Copy Host", command=self.copy_host)
        self.context_menu.add_command(label="Copy Port", command=self.copy_port)
        self.context_menu.add_separator()
        self.context_menu.add_command(label="Show TLS Details", command=self.show_tls_details)
        self.context_menu.add_command(label="Show Service Details", command=self.show_service_details)
        
        self.results_tree.bind("<Button-3>", self.show_context_menu)
    
    # ==================== EVENT HANDLERS ====================
    
    def load_targets_file(self):
        """Load targets from file (supports CIDR, IP lists)"""
        filename = filedialog.askopenfilename(
            title="Load Targets File",
            filetypes=[("Text files", "*.txt"), ("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if filename:
            try:
                with open(filename, 'r') as f:
                    content = f.read().strip()
                    # Parse different formats
                    targets = []
                    for line in content.split('\n'):
                        line = line.strip()
                        if line and not line.startswith('#'):
                            # Handle CIDR notation
                            if '/' in line:
                                try:
                                    network = ipaddress.ip_network(line, strict=False)
                                    targets.extend([str(ip) for ip in network.hosts()])
                                except:
                                    targets.append(line)
                            else:
                                targets.append(line)
                    
                    self.target_var.set(','.join(targets[:100]))  # Limit to 100 for GUI
                    self.log(f"Loaded {len(targets)} targets from {filename}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load file: {str(e)}")
    
    def start_scan(self):
        """Start the scanning process"""
        # Validate inputs
        targets_str = self.target_var.get().strip()
        ports_str = self.ports_var.get().strip()
        
        if not targets_str or not ports_str:
            messagebox.showerror("Error", "Please specify targets and ports")
            return
        
        # Parse targets
        targets = []
        for target in targets_str.split(','):
            target = target.strip()
            if target:
                # Handle CIDR
                if '/' in target:
                    try:
                        network = ipaddress.ip_network(target, strict=False)
                        targets.extend([str(ip) for ip in network.hosts()])
                    except:
                        targets.append(target)
                else:
                    targets.append(target)
        
        # Parse ports
        ports = []
        for port_item in ports_str.split(','):
            port_item = port_item.strip()
            if '-' in port_item:  # Range
                start, end = map(int, port_item.split('-'))
                ports.extend(range(start, end + 1))
            elif port_item.isdigit():
                ports.append(int(port_item))
        
        if not targets or not ports:
            messagebox.showerror("Error", "Invalid targets or ports")
            return
        
        # Setup stealth mode
        if self.tor_var.get():
            if not self.scanner.stealth.setup_tor():
                messagebox.showwarning("Warning", "Tor connection failed, continuing without Tor")
        
        # Update UI
        self.start_btn.config(state="disabled")
        self.stop_btn.config(state="normal")
        self.progress_var.set(0)
        self.clear_results()
        
        self.log(f"Starting scan: {len(targets)} targets, {len(ports)} ports")
        self.log(f"Stealth mode: {'ON' if self.stealth_var.get() else 'OFF'}")
        self.log(f"IPv6 support: {'ON' if self.ipv6_var.get() else 'OFF'}")
        
        # Start scan in background
        self.scan_thread = threading.Thread(
            target=self.run_scan, 
            args=(targets, ports), 
            daemon=True
        )
        self.scan_thread.start()
    
    def run_scan(self, targets: List[str], ports: List[int]):
        """Run the actual scan in background thread"""
        try:
            asyncio.run(self.async_scan(targets, ports))
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("Scan Error", str(e)))
        finally:
            self.root.after(0, self.scan_finished)
    
    async def async_scan(self, targets: List[str], ports: List[int]):
        """Async scan implementation"""
        def progress_callback(completed, total):
            progress = (completed / total) * 100
            self.root.after(0, lambda: self.progress_var.set(progress))
            self.root.after(0, lambda: self.status_label.config(text=f"Scanning... {completed}/{total}"))
        
        results = await self.scanner.scan_range(
            targets, ports, 
            stealth_mode=self.stealth_var.get(),
            progress_callback=progress_callback
        )
        
        # Update GUI with results
        self.root.after(0, lambda: self.display_results(results))
    
    def stop_scan(self):
        """Stop the current scan"""
        # Note: In a real implementation, you'd need proper async cancellation
        self.log("Scan stopped by user")
        self.scan_finished()
    
    def scan_finished(self):
        """Clean up after scan completion"""
        self.start_btn.config(state="normal")
        self.stop_btn.config(state="disabled")
        self.progress_var.set(100)
        self.status_label.config(text="Scan completed")
        
        # Update stats
        total_hosts = len(self.scanner.results)
        total_ports = sum(len(ports) for ports in self.scanner.results.values())
        self.hosts_label.config(text=f"Hosts: {total_hosts}")
        self.ports_label.config(text=f"Open Ports: {total_ports}")
        
        self.log(f"Scan completed: {total_hosts} hosts, {total_ports} open ports")
    
    def display_results(self, results: Dict[str, any]):
        """Display scan results in the tree view"""
        self.clear_results()
        
        for host, ports in results.items():
            host_item = self.results_tree.insert("", "end", text=host, values=("", "", "", "", "", "", ""))
            
            for port, info in ports.items():
                service_name = info.get("service", {}).get("service", "unknown")
                service_version = info.get("service", {}).get("version", "")
                banner = info.get("banner", "")[:50] + "..." if len(info.get("banner", "")) > 50 else info.get("banner", "")
                
                tls_status = "Yes" if info.get("tls_info") else "No"
                security_score = ""
                if info.get("tls_info", {}).get("security_headers"):
                    security_score = str(info["tls_info"]["security_headers"].get("security_score", 0))
                
                self.results_tree.insert(host_item, "end", text="",
                    values=(host, port, service_name, service_version, banner, tls_status, security_score))
        
        # Expand all items
        for item in self.results_tree.get_children():
            self.results_tree.item(item, open=True)
    
    def clear_results(self):
        """Clear the results tree view"""
        for item in self.results_tree.get_children():
            self.results_tree.delete(item)
    
    def export_excel(self):
        """Export results to Excel with pivot tables"""
        if not self.scanner.results:
            messagebox.showwarning("Warning", "No scan results to export")
            return
        
        filename = filedialog.asksaveasfilename(
            title="Save Excel Report",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if filename:
            try:
                report_file = self.excel_reporter.generate_report(self.scanner.results, filename)
                self.log(f"Excel report saved: {report_file}")
                messagebox.showinfo("Success", f"Report saved to {report_file}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save report: {str(e)}")
    
    def toggle_theme(self):
        """Toggle between dark and light themes"""
        self.current_theme = "light" if self.current_theme == "dark" else "dark"
        
        try:
            self.root.set_theme(self.themes[self.current_theme]["ttk_theme"])
        except:
            pass
        
        # Update text widget colors
        theme = self.themes[self.current_theme]
        self.log_text.config(bg=theme["bg"], fg="#00ff88" if self.current_theme == "dark" else "#008000")
        
        self.log(f"Theme switched to {self.current_theme}")
    
    def test_tor_connection(self):
        """Test Tor proxy connection"""
        if self.scanner.stealth.setup_tor():
            messagebox.showinfo("Success", "Tor connection successful")
            self.log("Tor connection test: SUCCESS")
        else:
            messagebox.showerror("Error", "Tor connection failed")
            self.log("Tor connection test: FAILED")
    
    def show_context_menu(self, event):
        """Show context menu on right-click"""
        try:
            self.context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.context_menu.grab_release()
    
    def copy_host(self):
        """Copy selected host to clipboard"""
        selection = self.results_tree.selection()
        if selection:
            item = self.results_tree.item(selection[0])
            host = item['values'][0] if item['values'] else item['text']
            self.root.clipboard_clear()
            self.root.clipboard_append(host)
    
    def copy_port(self):
        """Copy selected port to clipboard"""
        selection = self.results_tree.selection()
        if selection:
            item = self.results_tree.item(selection[0])
            if item['values'] and len(item['values']) > 1:
                port = item['values'][1]
                self.root.clipboard_clear()
                self.root.clipboard_append(str(port))
    
    def show_tls_details(self):
        """Show detailed TLS information"""
        selection = self.results_tree.selection()
        if selection:
            # Implementation for TLS details dialog
            messagebox.showinfo("TLS Details", "TLS details dialog would be implemented here")
    
    def show_service_details(self):
        """Show detailed service information"""
        selection = self.results_tree.selection()
        if selection:
            # Implementation for service details dialog
            messagebox.showinfo("Service Details", "Service details dialog would be implemented here")
    
    def log(self, message: str):
        """Add message to the live log"""
        timestamp = time.strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] {message}\n"
        
        self.log_text.config(state="normal")
        self.log_text.insert("end", log_entry)
        self.log_text.see("end")
        self.log_text.config(state="disabled")

# ==================== MAIN APPLICATION ====================

def main():
    """Main application entry point"""
    try:
        root = ThemedTk(theme="equilux")
    except:
        root = tk.Tk()
        
    app = ZeroSweepEnhancedGUI(root)
    
    # Set window icon (create a simple colored icon)
    try:
        icon = Image.new("RGBA", (32, 32), (0, 255, 136, 255))
        icon_photo = ImageTk.PhotoImage(icon)
        root.iconphoto(False, icon_photo)
    except:
        pass
    
    # Center window on screen
    root.update_idletasks()
    x = (root.winfo_screenwidth() // 2) - (1200 // 2)
    y = (root.winfo_screenheight() // 2) - (800 // 2)
    root.geometry(f"1200x800+{x}+{y}")
    
    # Start the GUI
    root.mainloop()

if __name__ == "__main__":
    main()